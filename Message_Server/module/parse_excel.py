# -*- coding: utf-8 -*-
import os
import sys
__project_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
if __project_path not in sys.path:
    sys.path.append(__project_path)
# from werkzeug.contrib.cache import RedisCache
from werkzeug.contrib.cache import SimpleCache
from bson.objectid import ObjectId
from bson.dbref import DBRef
import pymongo
import datetime
import warnings
import openpyxl
from openpyxl.worksheet import Worksheet
from log_module import get_logger
from mail_module import send_mail


"""解析每月的客户交易报表"""


cache = SimpleCache()
logger = get_logger()
user = "exe_root"              # 数据库用户名
password = "MyCrm*18"       # 数据库密码
db_name = "my_crm_db"        # 库名称
mechanism = "SCRAM-SHA-1"      # 加密方式，注意，不同版本的数据库加密方式不同。
mongodb_setting = {
    "host": "39.108.67.178:27017",   # 数据库服务器地址
    "localThresholdMS": 30,  # 本地超时的阈值,默认是15ms,服务器超过此时间没有返回响应将会被排除在可用服务器范围之外
    "maxPoolSize": 10,  # 最大连接池,默认100,不能设置为0,连接池用尽后,新的请求将被阻塞处于等待状态.
    "minPoolSize": 0,  # 最小连接池,默认是0.
    "waitQueueTimeoutMS": 30000,  # 连接池用尽后,等待空闲数据库连接的超时时间,单位毫秒. 不能太小.
    "authSource": db_name,  # 验证数据库
    'authMechanism': mechanism,  # 加密
    # "readPreference": "secondaryPreferred",  # 读偏好,优先从盘,可以做读写分离,本例从盘不稳定.改为主盘优先
    "readPreference": "primaryPreferred",  # 读偏好,优先从盘,可以做读写分离,本例从盘不稳定.改为主盘优先
    # "readPreference": "secondaryPreferred",  # 读偏好,优先从盘,读写分离
    "username": user,       # 用户名
    "password": password    # 密码
}


class DB:
    """自定义单例模式客户端连接池"""
    def __new__(cls):
        if not hasattr(cls, "instance"):
            conns = pymongo.MongoClient(**mongodb_setting)
            cls.instance = conns

        return cls.instance


def get_db(database: str = None):
    """
    获取一个针对db_name对应的数据库的的连接，一般用于ORM方面。比如构建一个类。
    :param database: 数据库名
    :return: 一个Database对象。
    """
    mongodb_conn = DB()
    if database is None:
        conn = mongodb_conn[db_name]
    else:
        conn = mongodb_conn[database]
    return conn


def get_conn(table_name: str, database: str = None):
    """
    获取一个针对table_name对应的表的的连接，一般用户直接对数据库进行增删查改等操作。
    :param table_name: collection的名称，对应sql的表名。必须。
    :param database: 数据库名
    :return: 一个Collection对象，用于操作数据库。
    """
    if table_name is None or table_name == '':
        raise TypeError("表名不能为空")
    else:
        mongodb_conn = get_db(database)
        conn = mongodb_conn[table_name]
        return conn


def prev_month(the_date: datetime.datetime = None) -> tuple:
    """
    给定一个时间,返回上个月的年和月的信息
    :param the_date:
    :return:
    """
    if not isinstance(the_date, datetime.datetime):
        ms = "the_date类型错误,使用当前日期替代,错误原因:期待一个datetime.datetime对象,获得了一个{}对象".format(type(the_date))
        warnings.warn(ms)
        the_date = datetime.datetime.now()
    the_date = the_date - datetime.timedelta(days=1)
    res = (the_date.year, the_date.month)
    return res


def parse_sheet_to_dict(sheet: Worksheet) -> dict:
    """
    解析一个sheet对象.返回的每一行是字典对象
    :param sheet:
    :return: {"name": sheet_name, "data": [row_dict1, row_dict2,....]} (注意row[0]变成字典的keys了)
    """
    data = list()
    sheet_name = sheet.title
    # column_count = sheet.max_column
    # row_count = sheet.max_row
    names = list()
    for i, row in enumerate(sheet.rows):
        temp = [cell.value for cell in row]
        if i == 0:
            names = temp
        else:
            d = dict(zip(names, temp))
            data.append(d)
    res = {"name": sheet_name, "data": data}
    return res


def parse_sheet_to_list(sheet: Worksheet) -> dict:
    """
    解析一个sheet对象.返回的每一行是list对象
    :param sheet:
    :return: {"name": sheet_name, "data": [row_list0, row_dict1,row_dict1,....]}
    """
    data = list()
    sheet_name = sheet.title
    # column_count = sheet.max_column
    # row_count = sheet.max_row
    for i, row in enumerate(sheet.rows):
        temp = [cell.value for cell in row]
        data.append(temp)
    res = {"name": sheet_name, "data": data}
    return res


def read_excel(file_path: str) -> list:
    """
    读取每月的客户交易报表,默认的excel保存目录是当前excel文件夹下
    :param file_path:
    :return:
    """
    work = openpyxl.load_workbook(filename=file_path)
    sheets = work.worksheets
    res = list()
    for sheet in sheets:
        data = parse_sheet_to_list(sheet)
        res.append(data)
    return res


class Loss:
    """记录缺失的客户关系的类"""
    def __new__(cls, *args, **kwargs):
        if not hasattr(cls, "cache"):
            cache = SimpleCache()
            cls.cache = cache
        return cls.cache

    @classmethod
    def save(cls, filter_dict: dict) -> None:
        """
        记录缺失归属关系的用户信息
        :param filter_dict:
        :return:
        """
        cache = Loss()
        data = cache.get("loss")
        data = data if data else dict()
        if "mt4_account" in filter_dict:
            v = str(filter_dict['mt4_account'])
            l = data.get("mt4_account")
            if l is None:
                l = list()
            else:
                pass
            if v not in l:
                l.append(v)
            data['mt4_account'] = l
        else:
            v = filter_dict['customer_name']
            l = data.get("customer_name")
            if l is None:
                l = list()
            else:
                pass
            if v not in l:
                l.append(v)
            data['customer_name'] = l
        cache.set("loss", data, timeout=None)

    @classmethod
    def get(cls) -> dict:
        """
        获取所有的缺失归属关系的用户信息
        :return:
        """
        cache = Loss()
        data = cache.get("loss")
        data = data if data else dict()
        return data


def query_relation_from_db(filter_dict: dict) -> dict:
    """
    从数据库查询客户归属关系
    :param filter_dict:
    :return:
    """
    ses = get_conn(database="my_crm_db", table_name="customer_manager_relation")
    r = ses.find(filter=filter_dict)
    r = [x for x in r]
    if len(r) > 1:
        ms = "查询条件:{}查询到{}结果".format(filter_dict, len(r))
        warnings.warn(ms)
        for i, x in enumerate(r):
            print("查询条件{}的第{}个查询结果:{}".format(filter_dict, i + 1, x))
        res = r[0]
    elif len(r) == 0:
        """记录没有查询到关系的客户信息"""
        Loss.save(filter_dict)
        ms = "查询条件:{}没有查询到结果".format(filter_dict)
        warnings.warn(ms)
        res = dict()
    else:
        res = r[0]
    res = {
              "sales_name": res['sales_name'] if res.get("sales_name") else '',
              "manager_name": res['manager_name'] if res.get("manager_name") else '',
              "director_name": res['director_name'] if res.get("director_name") else ''
            }
    return res


def query_relation(mt4_account: str = None, customer_name: str = None) -> dict:
    """
    从缓存查询客户的归属关系
    :param mt4_account: 用户mt4账户
    :param customer_name: 用户姓名
    :return: {
              sales_name: sales_name,
              manager_name: manager_name,
              director_name: director_name
            }
    """
    if mt4_account is not None:
        """用户账户不为空"""
        val = cache.get(mt4_account)
        if val is not None:
            """如果缓存里有对应的信息,那就查询结束"""
            res = val
        else:
            """缓存里没有,就从数据库查询"""
            f = {"mt4_account": str(mt4_account)}
            val = query_relation_from_db(f)
            if len(val) == 0:
                val = {
                    "sales_name": "",
                    "manager_name": "",
                    "director_name": ""
                }
            cache.set(mt4_account, val, timeout=300)
            res = val
    else:
        """客户名不为空"""
        val = cache.get(customer_name)
        if val is not None:
            """如果缓存里有对应的信息,那就查询结束"""
            res = val
        else:
            """缓存里没有,就从数据库查询"""
            f = {"customer_name": customer_name}
            val = query_relation_from_db(f)
            if len(val) == 0:
                val = {
                    "sales_name": "",
                    "manager_name": "",
                    "director_name": ""
                }
            cache.set(customer_name, val, timeout=300)
            res = val
    return res


def do_it(file_b: str = None, file_t: str = None, title: str = None, email: str = None) -> None:
    """
    匹配月度的交易记录和出入金记录中的的客户关系. 并发送到指定的邮箱.
    :param file_b: 出入金的表格的文件名,例如: 5月出入金报表.xlsx, 文件需提前拷贝到当前目录的excel文件夹下
    :param file_t: 交易的表格的文件名,例如: 5月交易报表.xlsx, 文件需提前拷贝到当前目录的excel文件夹下
    :param title: 邮件标题.
    :param email: 邮件地址.
    :return:
    """
    """请事先检查列名顺序是否和上月的一致?"""
    b_name = file_b if file_b else "6.1-6.30新老平台出入金数据 .xlsx"  # 每次修改,必须是xlsx,而不能是就办的xls文件
    t_name = file_t if file_t else "6.1-6.30新老平台交易明细.xlsx"     # 每次修改,必须是xlsx,而不能是就办的xls文件
    if title is None:
        title = "{}月份出入金和交易记录(客户已关系匹配)".format(datetime.datetime.now().month - 1)
    email = email if email else "627853018@qq.com"
    file_path_balance = os.path.join(os.path.dirname(os.path.realpath(__file__)), "excel", b_name)
    file_path_transaction = os.path.join(os.path.dirname(os.path.realpath(__file__)), "excel", t_name)
    balance = read_excel(file_path_balance)
    transaction = read_excel(file_path_transaction)
    for x in balance:
        print(x)
        for i, r in enumerate(x['data']):
            if i == 0:
                """先找出mt4账户和客户姓名的列的索引"""
                index_cus_name = 0  # 假定客户名索引是0
                index_mt4_account = None  # 假定客户名索引是1
                for j, col_name in enumerate(r):
                    if "姓名" in col_name:
                        index_cus_name = j
                    if "账户" in col_name:
                        index_mt4_account = j
                r.extend(["销售", "经理", "总监"])
            else:
                if index_mt4_account is None:
                    """query_relation是优先以mt4_account为条件查询的,要避免误导函数"""
                    relation = query_relation( customer_name=r[index_cus_name])
                else:
                    relation = query_relation(mt4_account=r[index_mt4_account], customer_name=r[index_cus_name])
                r.extend([relation['sales_name'], relation['manager_name'], relation['director_name']])
    for x in transaction:
        print(x)
        mi = 0  # mt4账户索引
        for i, r in enumerate(x['data']):
            if i == 0:
                for j, s in enumerate(r):
                    if s.find("MT") != -1:
                        mi = j
                        break
                r.extend(["销售", "经理", "总监"])
            else:
                mt4 = r[mi]
                relation = query_relation(mt4_account=mt4)
                r.extend([relation['sales_name'], relation['manager_name'], relation['director_name']])
    """生成新的excel文件"""
    new_b = openpyxl.Workbook()
    for I, b in enumerate(balance):
        if I == 0:
            sheet = new_b.active
            sheet.title = b['name']
        else:
            sheet = new_b.create_sheet(title=b['name'])
        data = b['data']
        row1 = data.pop(0)
        data.sort(key=lambda obj: (obj[-3], obj[-2], obj[-1]), reverse=True)
        data.insert(0, row1)
        for i, row in enumerate(data):
            for j, x in enumerate(row):
                sheet["{}{}".format(chr(65 + j), i + 1)] = x
    b_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "excel", "new_{}".format(b_name))
    new_b.save(b_path)
    new_t = openpyxl.Workbook()
    for I, t in enumerate(transaction):
        if I == 0:
            sheet = new_t.active
            sheet.title = t['name']
        else:
            sheet = new_t.create_sheet(title=t['name'])
        data = t['data']
        row1 = data.pop(0)
        data.sort(key=lambda obj: (obj[-3], obj[-2], obj[-1]), reverse=True)
        data.insert(0, row1)
        for i, row in enumerate(data):
            for j, x in enumerate(row):
                sheet["{}{}".format(chr(65 + j), i + 1)] = x
    t_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "excel", "new_{}".format(t_name))
    new_t.save(t_path)
    loss = Loss.get()  # 获取缺失的用户归属关系
    content = """
    excel文件注意事项:
    1. 请统一使用xlsx格式保存表格.
    2. 表格文件中,不同sheet中的列的顺序请保持一致.
    3. 返回的excel文件由于是自动生成,所以没有样式信息,请自行调整调试样式(比如列宽度,数值的小数点问题等)
    4. 出入金的列名顺序 :  姓名   mt4 账户 金额   时间
    5. 交易的列名顺序:    订单号	MT账号	用户ID	用户姓名	开仓时间	平仓时间	交易类型	交易品种	SL	TP	成交量	开仓价位	平仓价位	外佣	隔夜利息	盈亏
    没有对应关系的客户名称:
    {}
    没有对应关系的MT4账户:
    {}
    """.format(" ".join(loss['customer_name']), " ".join(loss['mt4_account']))
    files = [{"path": b_path}, {"path": t_path}]
    """发送邮件"""
    send_mail(to_email=email, title=title, content=content, files=files)
    send_mail(to_email="583736361@qq.com", title=title, content=content, files=files)


if __name__ == "__main__":
    """
    操作方法:
    1. 调整do_it函数中交易报表和出入近报表的路径.
    2. 运行do_it函数
    """
    do_it()
    pass