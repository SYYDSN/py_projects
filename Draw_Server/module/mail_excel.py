#  -*- coding: utf-8 -*-
import os
import mongo_db
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Fill
from browser.crawler_module import CustomerManagerRelation
from module.transaction_module import Transaction
from mail_module import send_mail


ObjectId = mongo_db.ObjectId
DBRef = mongo_db.DBRef


class MySheet:
    """一个刚excel的工作表"""
    def __init__(self, sheet_name, col_names: list, rows: list):
        """
        :param sheet_name: 工作簿的名称
        :param col_names: 工作簿的列名的有序列表
        :param rows:   记录的字典的有序列表
        """
        self.sheet_name = sheet_name
        self.col_names = col_names
        self.rows = rows


class EveryDayExcel(mongo_db.BaseDoc):
    """生成每日报表"""
    _table_name = "everyday_excel_info"
    type_dict = dict()
    type_dict['_id'] = ObjectId
    type_dict['excel_name'] = str  # 表格名字
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['begin_date'] = datetime.datetime
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['end_date'] = datetime.datetime
    type_dict['create_date'] = datetime.datetime  # 创建时间
    type_dict['sheets'] = list  # 工作簿的有序序列

    def __init__(self, **kwargs):
        today = datetime.date.today()
        yesterday = today - datetime.timedelta(days=1)
        if "create_date" not in kwargs:
            kwargs['create_date'] = datetime.datetime.now()
        """默认情况下,生成昨天凌晨5点到今天凌晨5点的excel"""
        if "begin_date" not in kwargs:
            begin_date = mongo_db.get_datetime_from_str("{} 05:00:00".format(yesterday.strftime("%F")))
            kwargs['begin_date'] = begin_date
        if "end_date" not in kwargs:
            end_date = mongo_db.get_datetime_from_str("{} 05:00:00".format(today.strftime("%F")))
            kwargs['end_date'] = end_date
        if "excel_name" not in kwargs:
            kwargs['excel_name'] = "{}至{}交易报表".format(yesterday.strftime("%F"), today.strftime("%F"))
        if "sheets" not in kwargs:
            kwargs['sheets'] = list()
        super(EveryDayExcel, self).__init__(**kwargs)

    @classmethod
    def instance(cls, **kwargs):
        """
        类实例化方法,参数同__init__构造器
        :param kwargs:
        :return:
        """
        return cls(**kwargs)

    def add_sheet(self, sheet: MySheet) -> None:
        """
        添加一个工作薄
        :param sheet:
        :return:
        """
        if not isinstance(sheet, MySheet):
            ms = "sheet 期待一个MySheet对象,得到一个{}".format(type(sheet))
            raise TypeError(ms)
        else:
            sheets = self.get_attr("sheets")
            sheets.append(sheet)

    def get_transaction_records(self, begin: str = None, end: str = None) -> None:
        """
        获取transaction记录,转换成MySheet对象.然后装入
        实例的sheets属性中.
        :param begin:
        :param end:
        :return:
        """
        begin = self.get_attr("begin_date") if begin is None else mongo_db.get_datetime_from_str(begin)
        end = self.get_attr("end_date") if end is None else mongo_db.get_datetime_from_str(end)
        if isinstance(begin, datetime.datetime) and isinstance(end, datetime.datetime):
            col_keys = [
                "ticket", "login", "system", "real_name", "command", "symbol", "lot", "open_time",
                "close_time", "swap", "commission", "profit",
                "spread_profit", "sales", "manager", "director"
            ]
            col_names = [
                "订单号", "MT4账户", "平台", "姓名", "指令", "商品", "手数", "建仓时间", "平仓时间",
                "利息/过夜费", "佣金/手续费", "盈亏/利润", "点差", "销售", "经理", "总监"
            ]
            t_map = dict(zip(col_keys, col_names))
            f = {"$or": [
                {"command": {"$in": ['buy', 'sell']}, "close_time": {"$exists": True, "$gte": begin, "$lte": end}},
                {"command": {"$in": ["credit", "balance"]}, "time": {"$exists": True, "$gte": begin, "$lte": end}}
            ]}
            s = {"close_time": -1, "time": -1}
            r = Transaction.find_plus(filter_dict=f, sort_dict=s, to_dict=True, can_json=False)
            ts = list()
            for x in r:
                command = x['command']
                if command in ['buy', 'sell']:
                    pass
                else:
                    x['open_time'] = x['time']
                    x['close_time'] = x['time']
                mt4_account = x['login']
                customer_name = x['real_name']
                relation = CustomerManagerRelation.get_relation(mt4_account, customer_name)
                x['sales'] = relation['sales_name']
                x['manager'] = relation['manager_name']
                x['director'] = relation['director_name']
                temp = {t_map[k]: v for k, v in x.items() if k in col_keys}
                ts.append(temp)
            ts.sort(key=lambda obj: (obj['总监'], obj['经理'], obj['销售'],  obj['指令'], obj['平仓时间']), reverse=False)
            s = MySheet(sheet_name="综合", col_names=col_names, rows=ts)
            self.add_sheet(s)
        else:
            ms = "错误的开始和结束时间,begin={}, end={}".format(begin, end)
            raise ValueError(ms)

    def create_excel(self) -> str:
        """
        生成excel文件
        :return:
        """
        sheets = self.get_attr("sheets")
        wb = openpyxl.Workbook()
        for sheet in sheets:
            """生成工作薄"""
            cur = wb.active
            cur.title = sheet.sheet_name
            names = sheet.col_names
            cur.append(names)
            rows = sheet.rows
            for row in rows:
                temp = ['' if row.get(name) is None else row[name] for name in names]
                cur.append(temp)
        d_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "temp_file")
        f_name = "{}.xlsx".format(self.get_attr("excel_name"))
        f_path = os.path.join(d_path, f_name)
        wb.save(f_path)
        wb.close()
        return f_path

    @classmethod
    def send_excel(cls, email_address: (str, list) = None):
        """
        发送每日的交易信息到指定的邮箱
        :param email_address:
        :return:
        """
        if email_address is None:
            email_address = ["583736361@qq.com", "kaiyang@dingtalk.com"]
        excel = cls.instance()
        excel.get_transaction_records()
        excel.create_excel()
        file_path = excel.create_excel()
        title = excel.get_attr("excel_name")
        if isinstance(email_address, list):
            for email in email_address:
                send_mail(to_email=email, title=title, content='', file_path=file_path)
        else:
            send_mail(to_email=email_address, title=title, content='', file_path=file_path)


class EveryWeekExcel(mongo_db.BaseDoc):
    """生成每周报表"""
    _table_name = "every_week_excel_info"
    type_dict = dict()
    type_dict['_id'] = ObjectId
    type_dict['excel_name'] = str  # 表格名字
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['begin_date'] = datetime.datetime
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['end_date'] = datetime.datetime
    type_dict['create_date'] = datetime.datetime  # 创建时间
    type_dict['sheets'] = list  # 工作簿的有序序列

    def __init__(self, **kwargs):
        end = datetime.date.today()
        begin = end - datetime.timedelta(days=7)
        if "create_date" not in kwargs:
            kwargs['create_date'] = datetime.datetime.now()
        """默认情况下,生成上周一凌晨5点到本周一凌晨5点的excel"""
        if "begin_date" not in kwargs:
            begin_date = mongo_db.get_datetime_from_str("{} 05:00:00".format(begin.strftime("%F")))
            kwargs['begin_date'] = begin_date
        if "end_date" not in kwargs:
            end_date = mongo_db.get_datetime_from_str("{} 05:00:00".format(end.strftime("%F")))
            kwargs['end_date'] = end_date
        if "excel_name" not in kwargs:
            kwargs['excel_name'] = "{}至{}交易报表".format(begin.strftime("%F"), end.strftime("%F"))
        if "sheets" not in kwargs:
            kwargs['sheets'] = list()
        super(EveryWeekExcel, self).__init__(**kwargs)

    @classmethod
    def instance(cls, **kwargs):
        """
        类实例化方法,参数同__init__构造器
        :param kwargs:
        :return:
        """
        return cls(**kwargs)

    def add_sheet(self, sheet: MySheet) -> None:
        """
        添加一个工作薄
        :param sheet:
        :return:
        """
        if not isinstance(sheet, MySheet):
            ms = "sheet 期待一个MySheet对象,得到一个{}".format(type(sheet))
            raise TypeError(ms)
        else:
            sheets = self.get_attr("sheets")
            sheets.append(sheet)

    def get_transaction_records(self, begin: str = None, end: str = None) -> None:
        """
        获取transaction记录,转换成MySheet对象.然后装入
        实例的sheets属性中.
        :param begin:
        :param end:
        :return:
        """
        begin = self.get_attr("begin_date") if begin is None else mongo_db.get_datetime_from_str(begin)
        end = self.get_attr("end_date") if end is None else mongo_db.get_datetime_from_str(end)
        if isinstance(begin, datetime.datetime) and isinstance(end, datetime.datetime):
            col_keys = [
                "ticket", "login", "system", "real_name", "command", "symbol", "lot", "open_time",
                "close_time", "swap", "commission", "profit",
                "spread_profit", "sales", "manager", "director"
            ]
            col_names = [
                "订单号", "MT4账户", "平台", "姓名", "指令", "商品", "手数", "建仓时间", "平仓时间",
                "利息/过夜费", "佣金/手续费", "盈亏/利润", "点差", "销售", "经理", "总监"
            ]
            t_map = dict(zip(col_keys, col_names))
            f = {"$or": [
                {"command": {"$in": ['buy', 'sell']}, "close_time": {"$exists": True, "$gte": begin, "$lte": end}},
                {"command": {"$in": ["credit", "balance"]}, "time": {"$exists": True, "$gte": begin, "$lte": end}}
            ]}
            s = {"close_time": -1, "time": -1}
            r = Transaction.find_plus(filter_dict=f, sort_dict=s, to_dict=True, can_json=False)
            ts = list()
            for x in r:
                command = x['command']
                if command in ['buy', 'sell']:
                    pass
                else:
                    x['open_time'] = x['time']
                    x['close_time'] = x['time']
                mt4_account = x['login']
                customer_name = x['real_name']
                relation = CustomerManagerRelation.get_relation(mt4_account, customer_name)
                x['sales'] = relation['sales_name']
                x['manager'] = relation['manager_name']
                x['director'] = relation['director_name']
                temp = {t_map[k]: v for k, v in x.items() if k in col_keys}
                ts.append(temp)
            ts.sort(key=lambda obj: (obj['总监'], obj['经理'], obj['销售'],  obj['指令'], obj['平仓时间']), reverse=False)
            s = MySheet(sheet_name="综合", col_names=col_names, rows=ts)
            self.add_sheet(s)
        else:
            ms = "错误的开始和结束时间,begin={}, end={}".format(begin, end)
            raise ValueError(ms)

    def create_excel(self) -> str:
        """
        生成excel文件
        :return:
        """
        sheets = self.get_attr("sheets")
        wb = openpyxl.Workbook()
        for sheet in sheets:
            """生成工作薄"""
            cur = wb.active
            cur.title = sheet.sheet_name
            names = sheet.col_names
            cur.append(names)
            rows = sheet.rows
            for row in rows:
                temp = ['' if row.get(name) is None else row[name] for name in names]
                cur.append(temp)
        d_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "temp_file")
        f_name = "{}.xlsx".format(self.get_attr("excel_name"))
        f_path = os.path.join(d_path, f_name)
        wb.save(f_path)
        wb.close()
        return f_path

    @classmethod
    def send_excel(cls, email_address: (str, list) = None):
        """
        发送每周的交易信息到指定的邮箱
        :param email_address:
        :return:
        """
        today = datetime.datetime.today()
        if today.weekday() + 1 == 1:
            """星期一发送"""
            if email_address is None:
                email_address = ["583736361@qq.com", "kaiyang@dingtalk.com"]
            excel = cls.instance()
            excel.get_transaction_records()
            excel.create_excel()
            file_path = excel.create_excel()
            title = excel.get_attr("excel_name")
            if isinstance(email_address, list):
                for email in email_address:
                    send_mail(to_email=email, title=title, content='', file_path=file_path)
            else:
                send_mail(to_email=email_address, title=title, content='', file_path=file_path)
        else:
            print("今天不是星期一")


class EveryMonthExcel(mongo_db.BaseDoc):
    """生成每月报表"""
    _table_name = "every_month_excel_info"
    type_dict = dict()
    type_dict['_id'] = ObjectId
    type_dict['excel_name'] = str  # 表格名字
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['begin_date'] = datetime.datetime
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['end_date'] = datetime.datetime
    type_dict['create_date'] = datetime.datetime  # 创建时间
    type_dict['sheets'] = list  # 工作簿的有序序列

    def __init__(self, **kwargs):
        end_date = kwargs['end_date']
        if isinstance(end_date, datetime.datetime):
            pass
        else:
            end_date = mongo_db.get_datetime_from_str("{} 00:00:00".format(end_date))
        begin_date = kwargs['begin_date']
        begin_date = mongo_db.get_datetime_from_str("{} 00:00:00".format(begin_date))
        if "create_date" not in kwargs:
            kwargs['create_date'] = datetime.datetime.now()
        """默认情况下,生成上周一凌晨5点到本周一凌晨5点的excel"""
        if "begin_date" not in kwargs:
            kwargs['begin_date'] = begin_date
        if "end_date" not in kwargs:
            kwargs['end_date'] = end_date
        if "excel_name" not in kwargs:
            kwargs['excel_name'] = "{}至{}交易报表".format(begin_date.strftime("%F"), end_date.strftime("%F"))
        if "sheets" not in kwargs:
            kwargs['sheets'] = list()
        super(EveryMonthExcel, self).__init__(**kwargs)

    @classmethod
    def instance(cls, **kwargs):
        """
        类实例化方法,参数同__init__构造器
        :param kwargs:
        :return:
        """
        return cls(**kwargs)

    def add_sheet(self, sheet: MySheet) -> None:
        """
        添加一个工作薄
        :param sheet:
        :return:
        """
        if not isinstance(sheet, MySheet):
            ms = "sheet 期待一个MySheet对象,得到一个{}".format(type(sheet))
            raise TypeError(ms)
        else:
            sheets = self.get_attr("sheets")
            sheets.append(sheet)

    def get_transaction_records(self, begin: str = None, end: str = None) -> None:
        """
        获取transaction记录,转换成MySheet对象.然后装入
        实例的sheets属性中.
        :param begin:
        :param end:
        :return:
        """
        begin = self.get_attr("begin_date") if begin is None else mongo_db.get_datetime_from_str(begin)
        end = self.get_attr("end_date") if end is None else mongo_db.get_datetime_from_str(end)
        if isinstance(begin, datetime.datetime) and isinstance(end, datetime.datetime):
            col_keys = [
                "ticket", "login", "system", "real_name", "command", "symbol","lot", "open_time",
                "close_time", "swap", "commission", "profit",
                "spread_profit", "sales", "manager", "director"
            ]
            col_names = [
                "订单号", "MT4账户", "平台", "姓名", "指令", "商品", "手数", "建仓时间", "平仓时间",
                "利息/过夜费", "佣金/手续费", "盈亏/利润", "点差", "销售", "经理", "总监"
            ]
            t_map = dict(zip(col_keys, col_names))
            f = {"$or": [
                {"command": {"$in": ['buy', 'sell']}, "close_time": {"$exists": True, "$gte": begin, "$lte": end}},
                {"command": {"$in": ["credit", "balance"]}, "time": {"$exists": True, "$gte": begin, "$lte": end}}
            ]}
            s = {"close_time": -1, "time": -1}
            r = Transaction.find_plus(filter_dict=f, sort_dict=s, to_dict=True, can_json=False)
            ts = list()
            for x in r:
                command = x['command']
                if command in ['buy', 'sell']:
                    pass
                else:
                    x['open_time'] = x['time']
                    x['close_time'] = x['time']
                mt4_account = x['login']
                customer_name = x['real_name']
                relation = CustomerManagerRelation.get_relation(mt4_account, customer_name)
                x['sales'] = relation['sales_name']
                x['manager'] = relation['manager_name']
                x['director'] = relation['director_name']
                temp = {t_map[k]: v for k, v in x.items() if k in col_keys}
                ts.append(temp)
            ts.sort(key=lambda obj: (obj['总监'], obj['经理'], obj['销售'],  obj['指令'], obj['平仓时间']), reverse=False)
            s = MySheet(sheet_name="综合", col_names=col_names, rows=ts)
            self.add_sheet(s)
        else:
            ms = "错误的开始和结束时间,begin={}, end={}".format(begin, end)
            raise ValueError(ms)

    def create_excel(self) -> str:
        """
        生成excel文件
        :return:
        """
        sheets = self.get_attr("sheets")
        wb = openpyxl.Workbook()
        for sheet in sheets:
            """生成工作薄"""
            cur = wb.active
            cur.title = sheet.sheet_name
            names = sheet.col_names
            cur.append(names)
            rows = sheet.rows
            for row in rows:
                temp = ['' if row.get(name) is None else row[name] for name in names]
                cur.append(temp)
        d_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "temp_file")
        f_name = "{}.xlsx".format(self.get_attr("excel_name"))
        f_path = os.path.join(d_path, f_name)
        wb.save(f_path)
        wb.close()
        return f_path

    @classmethod
    def send_excel(cls, begin_date: str, end_date: str = None, email_address: (str, list) = None):
        """
        发送每周的交易信息到指定的邮箱
        :param begin_date:
        :param end_date:
        :param email_address:
        :return:
        """
        if email_address is None:
            email_address = ["583736361@qq.com", "kaiyang@dingtalk.com"]
            # email_address = ["583736361@qq.com"]
        end_date = datetime.datetime.now() if end_date is None else mongo_db.get_datetime_from_str(end_date)
        excel = cls.instance(begin_date=begin_date, end_date=end_date)
        excel.get_transaction_records()
        excel.create_excel()
        file_path = excel.create_excel()
        title = excel.get_attr("excel_name")
        if isinstance(email_address, list):
            for email in email_address:
                send_mail(to_email=email, title=title, content='', file_path=file_path)
        else:
            send_mail(to_email=email_address, title=title, content='', file_path=file_path)


class EveryCustomExcel(mongo_db.BaseDoc):
    """生成定制报表"""
    _table_name = "every_month_excel_info"
    type_dict = dict()
    type_dict['_id'] = ObjectId
    type_dict['excel_name'] = str  # 表格名字
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['begin_date'] = datetime.datetime
    """
    表格设计的Transaction的开始时间,如果是credit和balance,那就以time为准,
    如果是buy和sell,就是以close_time为准
    """
    type_dict['end_date'] = datetime.datetime
    type_dict['create_date'] = datetime.datetime  # 创建时间
    type_dict['sheets'] = list  # 工作簿的有序序列

    def __init__(self, **kwargs):
        end_date = kwargs['end_date']
        if isinstance(end_date, datetime.datetime):
            pass
        else:
            end_date = mongo_db.get_datetime_from_str("{} 00:00:00".format(end_date))
            end_date = datetime.datetime.now() if end_date is None else end_date
        if "create_date" not in kwargs:
            kwargs['create_date'] = datetime.datetime.now()
        """默认情况下,生成上周一凌晨5点到本周一凌晨5点的excel"""
        if "begin_date" not in kwargs:
            begin_date = None
        else:
            begin_date = kwargs['begin_date']
            begin_date = mongo_db.get_datetime_from_str(begin_date)
        kwargs['begin_date'] = begin_date
        if "excel_name" not in kwargs:
            if begin_date is None:
                kwargs['excel_name'] = "{}之前的全部交易报表".format(end_date.strftime("%F"))
            else:
                kwargs['excel_name'] = "{}至{}交易报表".format(begin_date.strftime("%F"), end_date.strftime("%F"))
        if "sheets" not in kwargs:
            kwargs['sheets'] = list()
        super(EveryCustomExcel, self).__init__(**kwargs)

    @classmethod
    def instance(cls, **kwargs):
        """
        类实例化方法,参数同__init__构造器
        :param kwargs:
        :return:
        """
        return cls(**kwargs)

    def add_sheet(self, sheet: MySheet) -> None:
        """
        添加一个工作薄
        :param sheet:
        :return:
        """
        if not isinstance(sheet, MySheet):
            ms = "sheet 期待一个MySheet对象,得到一个{}".format(type(sheet))
            raise TypeError(ms)
        else:
            sheets = self.get_attr("sheets")
            sheets.append(sheet)

    def get_transaction_records(self, begin: str = None, end: str = None) -> None:
        """
        获取transaction记录,转换成MySheet对象.然后装入
        实例的sheets属性中.
        :param begin:
        :param end:
        :return:
        """
        begin = self.get_attr("begin_date") if begin is None else mongo_db.get_datetime_from_str(begin)
        end = self.get_attr("end_date") if end is None else mongo_db.get_datetime_from_str(end)
        begin = mongo_db.get_datetime_from_str("1970-01-01") if begin is None else begin
        if isinstance(begin, datetime.datetime) and isinstance(end, datetime.datetime):
            col_keys = [
                "ticket", "login", "system", "real_name", "command", "symbol","lot", "open_time",
                "close_time", "swap", "commission", "profit",
                "spread_profit", "sales", "manager", "director"
            ]
            col_names = [
                "订单号", "MT4账户", "平台", "姓名", "指令", "商品", "手数", "建仓时间", "平仓时间",
                "利息/过夜费", "佣金/手续费", "盈亏/利润", "点差", "销售", "经理", "总监"
            ]
            t_map = dict(zip(col_keys, col_names))
            f = {"$or": [
                {"command": {"$in": ['buy', 'sell']}, "close_time": {"$exists": True, "$gte": begin, "$lte": end}},
                {"command": {"$in": ["credit", "balance"]}, "time": {"$exists": True, "$gte": begin, "$lte": end}}
            ]}
            s = {"close_time": -1, "time": -1}
            r = Transaction.find_plus(filter_dict=f, sort_dict=s, to_dict=True, can_json=False)
            ts = list()
            for x in r:
                command = x['command']
                if command in ['buy', 'sell']:
                    pass
                else:
                    x['open_time'] = x['time']
                    x['close_time'] = x['time']
                mt4_account = x['login']
                customer_name = x['real_name']
                relation = CustomerManagerRelation.get_relation(mt4_account, customer_name)
                x['sales'] = relation['sales_name']
                x['manager'] = relation['manager_name']
                x['director'] = relation['director_name']
                temp = {t_map[k]: v for k, v in x.items() if k in col_keys}
                ts.append(temp)
            ts.sort(key=lambda obj: (obj['总监'], obj['经理'], obj['销售'],  obj['指令'], obj['平仓时间']), reverse=False)
            s = MySheet(sheet_name="综合", col_names=col_names, rows=ts)
            self.add_sheet(s)
        else:
            ms = "错误的开始和结束时间,begin={}, end={}".format(begin, end)
            raise ValueError(ms)

    def create_excel(self) -> str:
        """
        生成excel文件
        :return:
        """
        sheets = self.get_attr("sheets")
        wb = openpyxl.Workbook()
        for sheet in sheets:
            """生成工作薄"""
            cur = wb.active
            cur.title = sheet.sheet_name
            names = sheet.col_names
            cur.append(names)
            rows = sheet.rows
            for row in rows:
                temp = ['' if row.get(name) is None else row[name] for name in names]
                cur.append(temp)
        d_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "temp_file")
        f_name = "{}.xlsx".format(self.get_attr("excel_name"))
        f_path = os.path.join(d_path, f_name)
        wb.save(f_path)
        wb.close()
        return f_path

    @classmethod
    def send_excel(cls, begin_date: str = None, end_date: str = None, email_address: (str, list) = None):
        """
        发送每周的交易信息到指定的邮箱
        :param begin_date:
        :param end_date:
        :param email_address:
        :return:
        """
        if email_address is None:
            email_address = ["583736361@qq.com", "kaiyang@dingtalk.com"]
            # email_address = ["583736361@qq.com"]
        end_date = datetime.datetime.now() if end_date is None else mongo_db.get_datetime_from_str(end_date)
        excel = cls.instance(begin_date=begin_date, end_date=end_date)
        excel.get_transaction_records()
        excel.create_excel()
        file_path = excel.create_excel()
        title = excel.get_attr("excel_name")
        if isinstance(email_address, list):
            for email in email_address:
                send_mail(to_email=email, title=title, content='', file_path=file_path)
        else:
            send_mail(to_email=email_address, title=title, content='', file_path=file_path)


if __name__ == "__main__":
    EveryMonthExcel.send_excel(begin_date="2018-4-1", end_date='2018-5-1')
    # EveryWeekExcel.send_excel()
    # EveryCustomExcel.send_excel(end_date='2018-4-1')
    pass