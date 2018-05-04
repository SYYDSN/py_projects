#  -*- coding: utf-8 -*-
import os
import sys
__project_dir__ = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
if __project_dir__ not in sys.path:
    sys.path.append(__project_dir__)
import mongo_db
import datetime
import openpyxl


"""excel操作模块，用于导入数据"""


excel_dir = "excel"
excel_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), excel_dir)


def get_excel_path_list(dir_path: str = None) -> list:
    """
    根据目录获取excel文件的列表
    :param dir_path:
    :return:
    """
    if dir_path is None:
        dir_path = excel_path
    r = list()
    for name in os.listdir(dir_path):
        file_path = os.path.join(dir_path, name)
        if os.path.isfile(file_path) and (name.lower().endswith("xls") or name.lower().endswith("xlsx")):
            r.append(file_path)
        else:
            pass
    return r


def read_sheet_01(sh) -> list:
    """
    读取excel文件里的工作簿，这里代码要根据读取的文件不同自行该比那
    :param sh:
    :return:
    """
    res = list()
    now = datetime.datetime.now()
    for i, tr in enumerate(sh):
        if i == 0:
            pass
        else:
            p = tr[6].value
            if p == '盛汇中国':
                p = 'shengfxchina'
            elif p == 'fx888':
                p = 'shengfx888'
            elif p == 'fx china':
                p = 'shengfxchina'
            else:
                pass
            init = {
                "customer_name": tr[3].value,
                "mt4_account": tr[5].value,
                "platform": p,
                "sales_name": tr[7].value,
                "manager_name": tr[8].value,
                "director_name": tr[9].value,
                "create_date": now
            }
            res.append(init)
    return res


def read_excel(file_path: str):
    """
    读excel文件。
    :param file_path:
    :return:
    """
    excel = openpyxl.load_workbook(file_path)
    sheets = excel.sheetnames
    res = list()
    for sheet in sheets:
        sheet = excel.get_sheet_by_name(sheet)
        lines = read_sheet_01(sheet)
        res.extend(lines)
    return res


def calculate(p_path):
    excel = openpyxl.load_workbook(p_path)
    sheets = excel.sheetnames
    res = list()
    for sheet in sheets:
        sheet = excel.get_sheet_by_name(sheet)
        for i, tr in enumerate(sheet):
            if i == 0:
                pass
            else:
                init = {
                    "ticket": tr[0].value,
                    "m4_account": tr[1].value,
                    "platform": tr[2].value,
                    "customer_name": tr[3].value,
                    "command": tr[4].value,
                    "lot": tr[6].value,
                    "close_date": tr[8].value,
                    "swap": tr[9].value,
                    "commission": tr[10].value,
                    "profit": tr[11].value,
                    "spread_profit": tr[12].value,
                    "sales_name": tr[13].value,
                    "manager_name": tr[14].value,
                    "director_name": tr[15].value
                }
                res.append(init)
    res.sort(key=lambda obj: [obj['close_date']], reverse=True)
    """打印总手数"""
    lot_count = sum([x['lot'] for x in res if isinstance(x['lot'], (float, int))])
    print("总手数:{}".format(lot_count))
    return res


def calculate2(p_path):
    excel = openpyxl.load_workbook(p_path)
    sheets = excel.sheetnames
    res = list()
    for sheet in sheets:
        sheet = excel.get_sheet_by_name(sheet)
        for i, tr in enumerate(sheet):
            if i == 0:
                pass
            else:
                init = {
                    "ticket": int(tr[0].value) if tr[0].value.isdigit() else tr[0].value,
                    "m4_account": tr[1].value,
                    "command": tr[3].value,
                    "lot": float(tr[4].value) if tr[4].value is not None else 0,
                    "close_date": tr[7].value,
                    "swap": float(tr[10].value) if tr[10].value is not None else 0,
                    "commission": float(tr[11].value) if tr[11].value is not None else 0,
                    "profit": float(tr[9].value) if tr[9].value is not None else 0
                }
                if init['close_date'] is not None:
                    res.append(init)
    res.sort(key=lambda obj: [obj['close_date']], reverse=True)
    """打印总手数"""
    lot_count = sum([x['lot'] for x in res if isinstance(x['lot'], (float, int))])
    print("总手数:{}".format(lot_count))
    return res


if __name__ == "__main__":
    p1 = "/home/walle/work/projects/Draw_Server/module/temp_file/2018-04-01至2018-05-01交易报表.xlsx"
    p2 = "/home/walle/work/projects/Draw_Server/module/temp_file/交易记录2018-05-03 12_02_52.xlsx"
    r1 = calculate(p1)
    r2 = calculate2(p2)
    r1 = [x for x in r1 if x['platform'] == "office.shengfxchina.com:8443"]
    print(len(r1))
    print(len(r2))
    print(len(r1) - len(r2))
    r1_keys = [x['ticket'] for x in r1]
    aa = [x for x in r2 if x['ticket'] not in r1_keys]
    print(len(aa))
    res = dict()
    for x in aa:
        command = x['command']
        temp = res.get(command)
        if temp is None:
            temp = list()
        temp.append(x)
        res[command] = temp
    for k, v in res.items():
        ids = ",".join([str(x['ticket']) for x in v])
        print("{}类型,共计{}单, 订单号为:{}".format(k, len(v), ids))
    pass