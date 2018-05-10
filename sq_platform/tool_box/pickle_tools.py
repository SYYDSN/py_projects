# -*- coding: utf-8 -*-
import os
import sys
__project_dir = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
if __project_dir not in sys.path:
    sys.path.append(__project_dir)
import mongo_db
import json
import datetime
import openpyxl
import pickle
from api.data.item_module import User
from api.data.item_module import GPS
import pandas as pd
from pandas import Series
from pandas import DataFrame
from tool_box.draw_tools import draw_multiple_ploy_line
from log_module import get_logger


"""数据腌制工具"""


logger = get_logger()
ObjectId = mongo_db.ObjectId
DBRef = mongo_db.DBRef


def transform_type(in_dict: dict) -> dict:
    """
    转换doc的value的类型
    :param in_dict:
    :return:
    """
    res = dict()
    for k, v in in_dict.items():
        if isinstance(v, datetime.datetime):
            v = v.strftime("%Y-%m-%d %H:%M:%S.%f")
        elif isinstance(v, ObjectId):
            v = str(v)
        elif isinstance(v, DBRef):
            v = str(v.id)
        else:
            pass
        if k != "loc":
            if k == "pr":
                res['source'] = "gps"
            elif k == "be":
                res['declination'] = v
            else:
                res[k] = v
        else:
            pass
    return res


def rebuild_json():
    """从本机数据库创建json文件"""
    f = {"user_id": {"$exists": True, "$type": 3}, "pr": "gps"}
    ses = mongo_db.get_conn("gps_info")
    records = ses.find(filter=f)
    records = [transform_type(x) for x in records]
    d_path = os.path.join(__project_dir, "tool_box", "json")
    if not os.path.exists(d_path):
        os.makedirs(d_path)
    f_path = os.path.join(d_path, "all_gps.json")
    with open(f_path, mode="w", encoding="utf-8") as file:
        json.dump(records, file)


def read_json() -> list:
    """
    从本地读取json文件
    :return:
    """
    d_path = os.path.join(__project_dir, "tool_box", "json")
    f_path = os.path.join(d_path, "all_gps.json")
    with open(f_path, mode="r", encoding="utf-8") as file:
        data = json.load(file)
    data.sort(key=lambda obj: mongo_db.get_datetime_from_str(obj['time']), reverse=True)
    data_by_user = dict()
    count_by_user = dict()
    for x in data:
        u_id = x['user_id']
        if u_id not in data_by_user:
            data_by_user[u_id] = [x]
            count_by_user[u_id] = 1
        else:
            data_by_user[u_id].append(x)
            count = count_by_user[u_id]
            count += 1
            count_by_user[u_id] = count
    count_by_user = [{"id": k, "count": v} for k, v in count_by_user.items()]
    count_by_user.sort(key=lambda obj: obj['count'], reverse=True)
    ids = count_by_user[0: 3]
    dd_path = os.path.join(d_path, "single_by_user")
    if not os.path.exists(dd_path):
        os.makedirs(dd_path)
    ids = [x['id'] for x in ids]
    for i, id in enumerate(ids):
        name = "user_{}_gps.json".format(i + 1)
        v = data_by_user[id]
        with open(os.path.join(dd_path, name), mode="w", encoding="utf-8") as f:
            json.dump(v, f)


def record_digest():
    """获取文件摘要"""
    d_path = os.path.join(__project_dir, "tool_box", "json", "single_by_user")
    names = os.listdir(d_path)
    for name in names:
        p = os.path.join(d_path, name)
        with open(p, "r", encoding="utf-8") as f:
            data = json.load(f)
        date_list = list()
        for x in data:
            the_date = x['time'].split(" ")[0]
            if the_date not in date_list:
                date_list.append(the_date)
        print(name)
        print(data[0]['user_id'])
        print(len(data))
        print(len(date_list))
        print(date_list)


def draw_user_gps(phones: list = list(), day_range: int = 60) -> dict:
    """
    从数据库读取指定手机号码的用户的记录,保存到本地文件
    :param phones:
    :param day_range: 向前追溯的时间范围,默认是60天,如果是None表示不设时间限制.那样有可能会导致过大的数据.
    :return:
    """

    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "pkl")
    """
    file_path = os.path.join(dir_path, "华新中转场驾驶员使用保驾犬明细_2018_5_9.xlsx")
    excel = openpyxl.load_workbook(file_path)
    sheets = excel.sheetnames
    users = list()
    for sheet in sheets:
        sheet = excel.get_sheet_by_name(sheet)
        lines = list()
        for tr in sheet:
            first = tr[0].value
            if first is not None and (isinstance(first, int) or first.isdigit()):
                line = dict()
                line['job_num'] = str(tr[1].value)
                line['real_name'] = tr[2].value
                line['phone_num'] = str(tr[3].value)
                line['line'] = tr[4].value
                lines.append(line)
        users.extend(lines)
    users = {x['phone_num']: x for x in users}
    if len(phones) == 0:
        phones = list(users.keys())

    f = {"phone_num": {"$in": phones}}

    rs = User.find_plus(filter_dict=f, to_dict=True)
    user_dbref_list = list()
    phone_map = dict()
    for x in rs:
        user_id = x['_id']
        phone = x['phone_num']
        user_dbref = DBRef(collection=User.get_table_name(), database="platform_db", id=user_id)
        user_dbref_list.append(user_dbref)
        users[phone]['user_id'] = user_id
        phone_map[user_dbref] = phone

    f = {"pr": "gps", "user_id": {"$in": user_dbref_list}}
    """
    f = {"pr": "gps"}
    end = datetime.datetime.now()
    begin = end - datetime.timedelta(days=day_range) if isinstance(day_range, int) and day_range > 0 else None
    if begin is None:
        f['time'] = {"$lte": end}
    else:
        f['time'] = {"$lte": end, "$gte": begin}
    projection = ['altitude', 'latitude', 'longitude', 'speed', 'time', 'be', 'pr', 'user_id']
    gps_list = GPS.find_plus(filter_dict=f, projection=projection,  to_dict=True)
    f_path = os.path.join(dir_path, "data.pkl")
    file = open(f_path, "wb")
    pickle.dump(gps_list, file)
    file.close()


def read_pkl() -> list:
    """读取本地的全部的gps文件"""
    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "pkl")
    f_path = os.path.join(dir_path, "data.pkl")
    with open(f_path, "rb") as f:
        all_gps = pickle.load(f)
    return all_gps


def save_top_20():
    """保存gps数据最多的前20到文件"""
    d = dict()
    all_gps = read_pkl()
    for gps in all_gps:
        user_id = str(gps['user_id'].id)
        temp = d.get(user_id)
        if temp is None:
            temp = list()
        gps["user_id"] = user_id
        temp.append(gps)
        d[user_id] = temp
    ll = [{"user_id": k, "data": v} for k, v in d.items()]
    ll.sort(key=lambda obj: len(obj['data']), reverse=True)
    res = ll[0: 20]
    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "pkl")
    f_path = os.path.join(dir_path, "top_20.pkl")
    with open(f_path, "wb") as f:
        pickle.dump(res, f)


def save_top1():
    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "pkl")
    f_path = os.path.join(dir_path, "top_20.pkl")
    with open(f_path, "rb") as f:
        all_gps = pickle.load(f)
    f2_path = os.path.join(dir_path, "top_1.pkl")
    with open(f2_path, "wb") as f:
        pickle.dump(all_gps[0], f)


def calculate_acceleration(data_list):
    """
    计算加速度,公安部对公交车的急加速/急减速标准如下:
    1, 一级, 速度变化>=10km/h ,公交车站立人员即使抓扶也会造成扭伤或摔伤.
    1, 二级, 速度变化>=8km/h ,公交车站立人员抓扶会有明显倾斜感,未抓扶会造成扭伤或摔伤.
    1, 三级, 速度变化>=6km/h ,公交车站立人员未抓扶容易会造成扭伤或摔伤.
    持续时间上一般要求持续一段时间,防止异常数据造成的误报,
    """
    points = list()
    prev = None
    for x in data_list:
        if prev is None:
            prev = x
        else:
            """计算速度变化差"""
            speed_delta = (x['speed'] - prev['speed']) / ((x['time'] - prev['time']).total_seconds() / 3600)
            print(speed_delta)


def test_calculate():
    """测试算法"""
    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "pkl")
    f_path = os.path.join(dir_path, "top_1.pkl")
    with open(f_path, "rb") as f:
        gps_dict = pickle.load(f)
    gps_data = gps_dict['data']
    """
    {
        "title": "我是总标题",
        "data":[
            {
                "label": "label_01",   # 曲线的标签,用做图示
                "lw": 0,               # 图示的位置 参考lw参数说明
                "color": 'c',          # 折线的颜色 参考颜色说明
                "linestyle": '-',      # 折线的样式 参考折线样式说明, 默认是 '-'
                "marker": '.',         # 端点的样式 参考端点样式说明, 默认是 None
                "data_x": x_list,      # x轴数据
                "data_y": y_list       # y轴数据
            },
            ...
        ]
    }
    """
    x_data = list()
    y_data = list()
    [{y_data.append(x['speed'] * 3.6): x_data.append(x['time'])} for x in gps_data]
    the_data = {
        "title": "我是总标题",
        "data": [
            {
                "label": "label_01",   # 曲线的标签,用做图示
                "lw": 0,               # 图示的位置 参考lw参数说明
                "color": 'c',          # 折线的颜色 参考颜色说明
                "linestyle": '-',      # 折线的样式 参考折线样式说明, 默认是 '-'
                "marker": '.',         # 端点的样式 参考端点样式说明, 默认是 None
                "data_x": x_data,      # x轴数据
                "data_y": y_data       # y轴数据
            }
        ]
    }
    draw_multiple_ploy_line(the_data)
    # calculate_acceleration(gps_data)


def create_excel():
    dir_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "pkl")
    f_path = os.path.join(dir_path, "top_20.pkl")
    with open(f_path, "rb") as f:
        all_gps = pickle.load(f)

    names = [
        'user_id', 'speed', 'time', 'amap', 'latitude',
        'source', 'altitude', 'longitude', 'id'
    ]
    for gps in all_gps:
        gps_set = gps['data']
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "gps_list"
        sheet.append(names)
        for x in gps_set:
            temp = list()
            temp.append(x['user_id'])
            temp.append(x['speed'] * 3.6)
            temp.append(x['time'])
            temp.append("amap")
            temp.append(x['latitude'])
            temp.append("gps")
            temp.append(x['altitude'])
            temp.append(x['longitude'])
            temp.append(str(x['_id']))
            sheet.append(temp)
        dir2_path = os.path.join(os.path.dirname(os.path.realpath(__file__)), "excel")
        if not os.path.exists(dir2_path):
            os.makedirs(dir2_path)
        f2_path = os.path.join(dir2_path, "{}_gps.xlsx".format(gps['user_id']))
        wb.save(f2_path)


if __name__ == "__main__":
    # draw_user_gps(day_range=None)
    # read_pkl()
    # save_top_20()
    # save_top1()
    # test_calculate()
    create_excel()
    pass